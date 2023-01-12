import pandas as pd
import openpyxl
import glob as gb

plan = []
jundados = []
contador = 0

for planilhas in gb.glob(r"C:\Users\Leandro\Documents\python\ProjetoPrintBot\planilhas\*.xlsx"):
    plan = [planilhas]
    for i in plan:
        p = openpyxl.load_workbook(i)
        ws = p.active

        Ems = ws['F3'].value
        print(Ems)
        Emsres = ws['H3'].value
        print(Emsres)
        Sub = ws['D15'].value
        print(Sub)
        Subres = ws['G15'].value
        print(Subres)
        Ali = ws['D16'].value
        print(Ali)
        Alires = ws['G16'].value 
        print(Alires)
            
        guarDados = openpyxl.Workbook()
        gds = guarDados.active
        dados = [
        [Ems,Emsres],
        [Sub,Subres],
        [Ali,Alires],
        ]

        contador += 1
        if contador >= 1:
            jundados += dados
            for t in jundados:
                gds.append(t)
                       
print(gds)

guarDados.save(r"C:\Users\Leandro\Documents\python\ProjetoPrintBot\planilhas\resultado\TesteMulti.xlsx")
