#FEITO POR FABIO LEANDRO LOPES DA CUNHA
#CODIGO PARA JUNÇÃO DE DADOS DE PLANILHAS DIFERENTES EM UMA SÓ

#import OPENPYXL - Importação da biblioteca openpyxl
import openpyxl
import glob as gb
import pandas as pd
import os

#lista_planilha = []
#for plandados in gb.glob("C:/Users/Leandro/Documents/python/ProjetoPrintBot/planilhas/*xlsx"):
    #print(plandados)
    #lista_planilha.append(pd.read_excel(plandados))

#tabela = pd.concat(lista_planilha)

#tabela.to_excel("C:/Users/Leandro/Documents/python/ProjetoPrintBot/planilhas/TesteMulti.xlsx")

caminho = "C:/Users/Leandro/Documents/python/ProjetoPrintBot/planilhas/*xlsx"

lista = []

for file in os.listdir[caminho]:
    if file.endswith('.xlsx')


















'''    
for i in lista_planilha:
    #lista_planilha.append(pd.read_excel(plandados))
    #Abre a planilha deseja com os dados     
    #tabela = pd.concat[plandados]
    #print(tabela)

    wb = openpyxl.load_workbook(lista_planilha) 
    ws = wb.active #seleciona a aba ativa da planilha para ser utilizada, logo a primeira aba
    print(wb)

    
        #Salva os dados da cedula escolhida em variaveis 
    Ems = ws['F3'].value
    Emsres = ws['H3'].value
    Sub = ws['B18'].value
    Subres = ws['E18'].value
    Ali = ws['B19'].value
    Alires = ws['E19'].value 
        #Comp = ws['B6'].value
        #Compres = ws['C6'].value
        #Pcf = ws['B32'].value
        #Pcfres = ws['E32'].value
        #Pcn = ws['B33'].value
        #Pcnres = ws['E33'].value
        # DataImp = ws['B253'].value
        # DataImpres = ws['D253'].value

        #cria uma nova planilha
    novatab = openpyxl.Workbook()
    nvs = novatab.active

        #coloca os valores pedidos da planilha na nova planilha

    dados = [
            [Ems,Emsres],
            [Sub,Subres],
            [Ali,Alires]#,
            #[Comp,Compres],
            #[Pcf,Pcfres],
            #[Pcn,Pcnres],
            #[DataImp,DataImpres],
        ]

    for i in dados:
        nvs.append(i)

        nvs['F3'].number_format = 'dd/mm/yyyy' #formata o dado em dia, mês e ano com 4 digitos
        #nvs['B7'].number_format = 'dd/mm/yyyy'
    #novatab.save("C:/Users/Leandro/Documents/python/ProjetoPrintBot/planilhas/*xlsx")
            

    novaplanilhas = pd.concat(nvs)

        #salva os valores na nova planilha
            #tabelaUnica = novaplanilhas

    novaplanilhas.to_excel('C:/Users/A2248/Documents/Downloads/PastaTeste/Nova pasta/TesteMultiMulti.xlsx',index=False)

    #novatab.save("C:/Users/A2248/Documents/Downloads/PastaTeste/Nova pasta/TesteMulti.xlsx")'''